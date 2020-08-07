import openpyxl
from openpyxl.styles.colors import RED,GREEN,BLUE,YELLOW
wb = openpyxl.Workbook()
ws = wb.active
#自定义格式字符串还可以通过分号，为单元格可能出现的 4 种类型的数据设置不同的格式
#这 4 种类型分别是   正值;负值;零值;文本
#当单元格的值是 “正值” 时，自定义格式为 "[
# RED]+#,###.00"
#当单元格的值是 “负值” 时，自定义格式为 "[GREEN]-#,###.00"
ws['A1'].number_format = "[RED]+#,###.00;[GREEN]-#,###.00"
ws['A1'] = 99
#
ws['A2'].number_format = "[RED]+#,###.00;[GREEN]-#,###.00"
ws['A2'] = -99

ws['A3'].number_format = "[RED];[GREEN];[BLUE];[YELLOW]"
ws['A3'] = 0

ws['A4'].number_format = "[RED];[GREEN];[BLUE];[YELLOW]"
ws['A4'] = "FishC"
#还可以将取值范围和颜色规则搭配使用，比如 "[<60][RED]不及格;[>=60]及格"
ws['A5'].number_format = "[=1]男;[=0]女"
ws['A5'] = 0

ws['A6'].number_format = "[=1]男;[=0]女"
ws['A6'] = 1

ws['A7'].number_format = "[=1]男;[=0]女"
ws['A7'] = 2

ws['A8'].number_format = "[<60][RED]不及格;[>=60][GREEN]及格"
ws['A8'] = 58

ws['A9'].number_format = "[<60][RED]不及格;[>=60][GREEN]及格"
ws['A9'] = 68
wb.save(r"number.xlsx")