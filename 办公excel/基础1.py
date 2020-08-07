import openpyxl
import datetime
wb=openpyxl.Workbook()
ws=wb.active #获取工作表
print(ws.title)#表标题
ws['A1']=250
ws.append([1,2,3])
#自动会转换格式
ws['A3']=datetime.datetime.now()
wb.save(r'C:\Users\Administrator\Desktop\1.xlsx')
