import openpyxl
import datetime
wb=openpyxl.Workbook()
ws=wb.active
ws['A1']=88.8
ws['A1'].number_format='#,###.00鱼币'

ws['A2']=datetime.datetime.today()
ws['A2'].number_format='yyyy-mm-dd'

wb.save("test2.xlsx")