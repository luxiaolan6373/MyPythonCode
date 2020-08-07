import openpyxl
#打开表格
wb=openpyxl.load_workbook(r"C:\Users\Administrator\Desktop\top250.xlsx")
#遍历工作表名字
print(wb.sheetnames)
#读取指定的工作表
ws=wb['Sheet']
nws=wb.create_sheet(index=0,title="FishDEMO")
print(wb.sheetnames)
#直接定位单元格
c=ws['A2']
#横 ,竖,坐标
print(c.column,c.row,c.coordinate)
#读取单元格内容
print(ws['A4'].value)
#将数字转换成列符号
print(openpyxl.cell.cell.get_column_letter(496))
#将列符号转换成数字 这句报错
#print(openpyxl.cell.cell.get_column_index_from_string('JB'))
#选定范围打印
for each_movie in ws['A2':'B10']:
    #each_movie就是一行的数据
    for each_cell in each_movie:
        #each_cell就是每一列的打印出来
        print(each_cell.value,end='')
    print('')#换行
#遍历数据
for each_row in ws.rows:
    print(each_row[0].value,each_row[1].value,each_row[2].value)
#选定范围
for each_row in ws.iter_rows(min_row=2,min_col=1,max_row=4,max_col=2):
    print(each_row[0].value,each_row[1].value)
#拷贝工作表
news=wb.copy_worksheet(ws)
#保存工作表
wb.save(r"C:\Users\Administrator\Desktop\top250.xlsx")