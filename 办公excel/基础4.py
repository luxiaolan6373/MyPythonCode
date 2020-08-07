from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import GradientFill
from openpyxl.styles import Border,Side
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle
#创建一个工作簿对象
wb=Workbook()
#创建工作表
ws=wb.active

b2=ws['B2']
b2.value='FishC'
#创建字体为 加粗 颜色红色
bold_red_font=Font(bold=True,color='FF0000')
b2.font=bold_red_font
b3=ws['B3']
b3.value='FishC'
#创建字体为 字号16 斜线 斜体,颜色蓝色
italic_strike_16font=Font(size=16,italic=True,strike=True,color='0000FF')
b3.font=italic_strike_16font
#创建填充为 纯色 颜色黄色
yellow_fill=PatternFill(fill_type='solid',fgColor="FFFF00")
b2.fill=yellow_fill
#创建渐变填充 从红到绿
red_to_green_fill=GradientFill(type='linear',stop=('FF0000',"00FF00"))
b3.fill=red_to_green_fill
#创建边框对象
thin_side= Side(border_style='thin',color='000000')
double_side= Side(border_style='double',color='FF0000')
#绘制对角线,两条都为真
b2.border=Border(diagonal=thin_side,diagonalUp=True,diagonalDown=True)
b3.border=Border(left=double_side,top=double_side,right=double_side,bottom=double_side)

#合并单元格
ws.merge_cells(('A4:C4'))
ws['A4'].value='I love FishC.com'
#设置对齐方式为 居中
center_alinment= Alignment(horizontal='center',vertical="center")
ws['A4'].alignment=center_alinment
#创建一个命名空间
highlight=NamedStyle(name='highlight')
#设置它的风格
highlight.font=Font(bold=True,size=20,color="FF8323")
highlight.alignment=Alignment(horizontal='center',vertical='center')

#注册命名空间风格 更加方便一点这种方式
wb.add_named_style(highlight)

ws['A4'].style=highlight
ws['B5'].value='Love'
ws['B5'].style=highlight


wb.save(r"C:\Users\Administrator\Desktop\3.xlsx")